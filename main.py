from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from kconfiglib import Kconfig
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


WORKBOOK_DEFAULT = Path("xlsx") / "E01 CHCM-1C-2C1V(得邦)_config_Dataset_LEFT_V0.2.xlsx"
KCONFIG_FILE = Path("Kconfig")
KCONFIG_CONFIG_DEFAULT = Path(".config")
KCONFIG_WORKBOOK_SYMBOL = "CHCM_WORKBOOK_PATH"
OUTPUT_DIR = Path("output")
FIELD_KEYS = {
    "C": "config_word_0",
    "D": "value_1",
    "E": "value_2",
}
OUTPUT_VALUE_KEYS = {
    "config_word_0": "value_1",
    "value_1": "value_2",
    "value_2": "value_3",
}
FIELD_LABELS = {
    "config_word_0": "配置字0",
    "value_1": "配置值1",
    "value_2": "配置值2",
}
VALUE_LIKE_PATTERN = re.compile(
    r"^(?:/|others|reserved|inactive|"
    r"\d+(?:\.\d+)?(?:\s*\(default\))?|"
    r"\d+(?:\.\d+)?-\d+(?:\.\d+)?(?:\s*\(default\))?)$",
    re.IGNORECASE,
)
DEFAULT_MARKER_PATTERN = re.compile(r"\s*\(default\)\s*", re.IGNORECASE)
CH_CFG_IC_PATTERN = re.compile(r"^CV_IC\d+$")
ANIMATION_CHANNEL_PATTERN = re.compile(r"^(IC\d+)-CH(\d+)$")
SUPPORTED_SHEET_NAMES = (
    "HCM_PriLIN_Matrix",
    "CH_Cfg",
    "Animation_Cfg",
    "Motor_Cfg",
    "TI_sequential",
)


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r\n", "\n").strip()
    return text or None


def normalize_field_value(value: str | None) -> str | None:
    if value in {None, "/"}:
        return None
    return value


def clean_output_value(value: str) -> str:
    return DEFAULT_MARKER_PATTERN.sub("", value).strip()


def normalize_json_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value

    text = clean_output_value(str(value).replace("\r\n", "\n").strip())
    return text or None


def load_workbook_path_from_kconfig(config_path: Path) -> Path | None:
    if not KCONFIG_FILE.is_file():
        return None

    kconfig = Kconfig(str(KCONFIG_FILE), warn=False)
    if config_path.is_file():
        kconfig.load_config(str(config_path))

    symbol = kconfig.syms.get(KCONFIG_WORKBOOK_SYMBOL)
    if symbol is None:
        raise ValueError(f"Kconfig 中未定义 {KCONFIG_WORKBOOK_SYMBOL}")

    workbook_path = normalize_text(symbol.str_value)
    if not workbook_path:
        return None
    return Path(workbook_path)


def resolve_workbook_path(cli_workbook: Path | None, config_path: Path) -> Path:
    if cli_workbook is not None:
        return cli_workbook

    workbook_from_kconfig = load_workbook_path_from_kconfig(config_path)
    if workbook_from_kconfig is not None:
        return workbook_from_kconfig

    return WORKBOOK_DEFAULT


def has_cjk(text: str) -> bool:
    return any("\u4e00" <= ch <= "\u9fff" for ch in text)


def is_value_like(text: str | None) -> bool:
    if text is None:
        return False
    return bool(VALUE_LIKE_PATTERN.fullmatch(text))


def is_label_like(text: str | None) -> bool:
    if text is None:
        return False
    if is_value_like(text):
        return False
    if has_cjk(text):
        return True
    return bool(re.fullmatch(r"Value\d+", text, re.IGNORECASE))


def is_item_start(item_code: str | None) -> bool:
    return bool(item_code and item_code.isdigit())


def row_snapshot(worksheet: Worksheet, row_index: int) -> dict[str, Any]:
    return {
        "source_row": row_index,
        "item_code": normalize_text(worksheet[f"A{row_index}"].value),
        "item_name": normalize_text(worksheet[f"B{row_index}"].value),
        "config_word_0": normalize_field_value(normalize_text(worksheet[f"C{row_index}"].value)),
        "value_1": normalize_field_value(normalize_text(worksheet[f"D{row_index}"].value)),
        "value_2": normalize_field_value(normalize_text(worksheet[f"E{row_index}"].value)),
        "description": normalize_text(worksheet[f"F{row_index}"].value),
    }


def row_has_content(row: dict[str, Any]) -> bool:
    return any(row[key] is not None for key in ("item_code", "item_name", "config_word_0", "value_1", "value_2", "description"))


def row_has_label_fields(row: dict[str, Any]) -> bool:
    present = [row[key] for key in FIELD_LABELS if row[key] is not None]
    if not present:
        return False
    return all(is_label_like(value) for value in present)


def extract_values(row: dict[str, Any]) -> dict[str, str]:
    return {key: row[key] for key in FIELD_LABELS if row[key] is not None}


def build_field_labels(
    used_keys: set[str],
    explicit_labels: dict[str, str] | None = None,
) -> dict[str, str]:
    explicit_labels = explicit_labels or {}
    return {key: explicit_labels.get(key, FIELD_LABELS[key]) for key in FIELD_LABELS if key in used_keys}


def detail_row_payload(row: dict[str, Any]) -> dict[str, Any]:
    payload: dict[str, Any] = {"source_row": row["source_row"]}
    for key in FIELD_LABELS:
        if row[key] is not None:
            payload[key] = row[key]
    if row["description"] is not None:
        payload["description"] = row["description"]
    return payload


def detect_used_keys(*sections: Any) -> set[str]:
    used: set[str] = set()
    for section in sections:
        if isinstance(section, dict):
            for key in FIELD_LABELS:
                if section.get(key) is not None:
                    used.add(key)
        elif isinstance(section, list):
            for item in section:
                if not isinstance(item, dict):
                    continue
                for key in FIELD_LABELS:
                    if item.get(key) is not None:
                        used.add(key)
    return used


def explicit_labels_from_row(row: dict[str, Any] | None) -> dict[str, str]:
    if row is None:
        return {}
    labels: dict[str, str] = {}
    for key in FIELD_LABELS:
        value = row[key]
        if is_label_like(value):
            labels[key] = value
    return labels


def classify_detail_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if not rows:
        return [], []
    if all(row_has_label_fields(row) for row in rows):
        return [detail_row_payload(row) for row in rows], []
    return [], [detail_row_payload(row) for row in rows]


def parse_config_block(top_row: dict[str, Any], detail_rows: list[dict[str, Any]]) -> dict[str, Any]:
    remaining_rows = [row for row in detail_rows if row_has_content(row)]
    explicit_labels = explicit_labels_from_row(top_row)
    defaults: dict[str, str] = {}
    definitions: list[dict[str, Any]] = []
    options: list[dict[str, Any]] = []
    input_hint = top_row["description"]
    structure_type = "free_input"

    if explicit_labels:
        defaults_row = None
        if remaining_rows and remaining_rows[0]["item_code"] is None and remaining_rows[0]["item_name"] is None:
            defaults_row = remaining_rows.pop(0)
            defaults = extract_values(defaults_row)
        if defaults_row and defaults_row["description"]:
            input_hint = defaults_row["description"]

        definitions, options = classify_detail_rows(remaining_rows)
        structure_type = "mapping_table" if options else "free_input"
        if defaults:
            structure_type = "composite_select" if options else "composite_input"
    else:
        defaults = extract_values(top_row)
        definitions, options = classify_detail_rows(remaining_rows)
        if options:
            structure_type = "select"
        elif definitions:
            structure_type = "free_input"

        if definitions and not input_hint:
            definition_descriptions = [row.get("description") for row in definitions if row.get("description")]
            if definition_descriptions:
                input_hint = definition_descriptions[0]

    definition_labels = explicit_labels_from_row(remaining_rows[0]) if definitions and remaining_rows else {}
    used_keys = detect_used_keys(defaults, definitions, options)
    field_labels = build_field_labels(used_keys, explicit_labels or definition_labels)

    item = {
        "id": int(top_row["item_code"]),
        "name": top_row["item_name"],
        "structure_type": structure_type,
        "source_row_start": top_row["source_row"],
        "field_labels": field_labels,
    }

    if defaults:
        item["defaults"] = defaults
    if input_hint:
        item["input_hint"] = input_hint
    if definitions:
        item["definitions"] = definitions
    if options:
        item["options"] = options
    return item


def parse_hcm_prilin_matrix(worksheet: Worksheet) -> dict[str, Any]:
    rows = [row_snapshot(worksheet, row_index) for row_index in range(1, worksheet.max_row + 1)]
    rows = [row for row in rows if row_has_content(row)]

    start_indices = [index for index, row in enumerate(rows) if is_item_start(row["item_code"])]
    if not start_indices:
        raise ValueError("未找到配置项块。")

    header_row = rows[0]
    first_positive_index = next(
        (index for index, row in enumerate(rows) if is_item_start(row["item_code"]) and row["item_code"] != "0"),
        None,
    )
    global_note_row_numbers = {
        row["source_row"]
        for row in rows[1:first_positive_index]
        if row["item_code"] is None
        and row["item_name"] is None
        and all(row[key] is None for key in FIELD_LABELS)
        and row["description"] is not None
    }
    notes = [row["description"] for row in rows if row["source_row"] in global_note_row_numbers]

    items: list[dict[str, Any]] = []
    for position, start_index in enumerate(start_indices):
        end_index = start_indices[position + 1] if position + 1 < len(start_indices) else len(rows)
        block_rows = rows[start_index:end_index]
        detail_rows = [row for row in block_rows[1:] if row["source_row"] not in global_note_row_numbers]
        items.append(parse_config_block(block_rows[0], detail_rows))

    result = {
        "parser": "hcm_prilin_matrix",
        "sheet_name_raw": worksheet.title,
        "sheet_name": worksheet.title.strip(),
        "header": {
            "item_code": header_row["item_code"],
            "item_name": header_row["item_name"],
            **{FIELD_KEYS[column]: normalize_text(worksheet[f"{column}{header_row['source_row']}"].value) for column in FIELD_KEYS},
            "description": header_row["description"],
        },
        "items": items,
    }

    if notes:
        result["notes"] = notes
    return result


def parse_ch_cfg(worksheet: Worksheet) -> dict[str, Any]:
    header_row_index = next(
        (row_index for row_index in range(1, worksheet.max_row + 1) if normalize_text(worksheet[f"A{row_index}"].value) == "IC.NO"),
        None,
    )
    if header_row_index is None:
        raise ValueError("CH_Cfg 中未找到 IC.NO 表头。")

    channel_headers: list[tuple[int, str]] = []
    for column_index in range(2, worksheet.max_column + 1):
        channel_name = normalize_text(worksheet.cell(header_row_index, column_index).value)
        if channel_name is not None:
            channel_headers.append((column_index, channel_name))

    ics: list[dict[str, Any]] = []
    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        ic_name = normalize_text(worksheet.cell(row_index, 1).value)
        if not ic_name or not CH_CFG_IC_PATTERN.fullmatch(ic_name):
            continue

        channels: dict[str, Any] = {}
        for column_index, channel_name in channel_headers:
            value = normalize_json_scalar(worksheet.cell(row_index, column_index).value)
            if value is not None:
                channels[channel_name] = value

        ics.append(
            {
                "ic_id": len(ics),
                "source_row": row_index,
                "ic_name": ic_name,
                "channels": channels,
            }
        )

    config_header_row_index = next(
        (row_index for row_index in range(1, worksheet.max_row + 1) if normalize_text(worksheet[f"B{row_index}"].value) == "配置类型说明"),
        None,
    )
    config_type_descriptions: dict[str, str] = {}
    if config_header_row_index is not None:
        for row_index in range(config_header_row_index + 1, worksheet.max_row + 1):
            code = normalize_json_scalar(worksheet.cell(row_index, 2).value)
            description = normalize_text(worksheet.cell(row_index, 3).value)
            if code is None:
                if config_type_descriptions:
                    break
                continue
            if description is None:
                continue
            config_type_descriptions[str(code)] = description

    return {
        "parser": "ch_cfg",
        "sheet_name_raw": worksheet.title,
        "sheet_name": worksheet.title.strip(),
        "channel_headers": [channel_name for _, channel_name in channel_headers],
        "config_type_descriptions": config_type_descriptions,
        "ics": ics,
    }


def find_animation_channel_headers(
    worksheet: Worksheet,
    header_row_index: int,
    *,
    start_column: int = 5,
) -> list[tuple[int, str]]:
    channel_headers: list[tuple[int, str]] = []
    for column_index in range(start_column, worksheet.max_column + 1):
        channel_name = normalize_text(worksheet.cell(header_row_index, column_index).value)
        if channel_name and ANIMATION_CHANNEL_PATTERN.fullmatch(channel_name):
            channel_headers.append((column_index, channel_name))
    if not channel_headers:
        raise ValueError("未找到 IC 通道表头。")
    return channel_headers


def summarize_animation_channels(channel_headers: list[tuple[int, str]]) -> tuple[int, int]:
    ic_channel_counts: dict[str, int] = {}
    for _, channel_name in channel_headers:
        match = ANIMATION_CHANNEL_PATTERN.fullmatch(channel_name)
        if match is None:
            continue
        ic_name = match.group(1)
        ic_channel_counts[ic_name] = ic_channel_counts.get(ic_name, 0) + 1

    if not ic_channel_counts:
        raise ValueError("Animation_Cfg 中未找到有效的 IC 通道。")

    unique_channel_counts = set(ic_channel_counts.values())
    if len(unique_channel_counts) != 1:
        details = ", ".join(f"{ic_name}:{count}" for ic_name, count in ic_channel_counts.items())
        raise ValueError(f"Animation_Cfg 中各 IC 通道数不一致: {details}")

    return len(ic_channel_counts), next(iter(unique_channel_counts))


def classify_animation_kind(mode_name: str) -> str:
    normalized = mode_name.strip().lower()
    if re.search(r"\bunlock\b", normalized):
        return "unlock"
    if re.search(r"\block\b", normalized):
        return "lock"
    return "other"


def summarize_animation_kind_counts(animations: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "unlock_animation_count": sum(1 for animation in animations if animation["animation_kind"] == "unlock"),
        "lock_animation_count": sum(1 for animation in animations if animation["animation_kind"] == "lock"),
    }


def group_animations_by_kind(animations: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped = {
        "unlock": [],
        "lock": [],
        "other": [],
    }
    for animation in animations:
        grouped[animation["animation_kind"]].append(animation)
    return grouped


def read_animation_channels(
    worksheet: Worksheet,
    row_index: int,
    channel_headers: list[tuple[int, str]],
    *,
    omit_zero: bool = False,
) -> dict[str, Any]:
    channels: dict[str, Any] = {}
    for column_index, channel_name in channel_headers:
        value = normalize_json_scalar(worksheet.cell(row_index, column_index).value)
        if value is not None:
            if omit_zero and value == 0:
                continue
            channels[channel_name] = value
    return channels


def parse_animation_cfg(worksheet: Worksheet) -> dict[str, Any]:
    header_row_index = next(
        (
            row_index
            for row_index in range(1, worksheet.max_row + 1)
            if normalize_text(worksheet.cell(row_index, 2).value) == "MODE"
            and normalize_text(worksheet.cell(row_index, 3).value) == "Times(ms)"
            and normalize_text(worksheet.cell(row_index, 4).value) == "CHANNEL"
        ),
        None,
    )
    if header_row_index is None:
        raise ValueError("Animation_Cfg 中未找到 MODE/Times(ms)/CHANNEL 表头。")

    channel_headers = find_animation_channel_headers(worksheet, header_row_index)
    total_ic_count, channel_count_per_ic = summarize_animation_channels(channel_headers)
    header = {
        "mode": normalize_text(worksheet.cell(header_row_index, 2).value),
        "time_ms": normalize_text(worksheet.cell(header_row_index, 3).value),
        "channel_type": normalize_text(worksheet.cell(header_row_index, 4).value),
    }

    animations: list[dict[str, Any]] = []
    current_animation: dict[str, Any] | None = None

    def finalize_animation() -> None:
        nonlocal current_animation
        if current_animation is not None:
            animations.append(current_animation)
            current_animation = None

    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        mode_name = normalize_text(worksheet.cell(row_index, 2).value)
        time_ms = normalize_json_scalar(worksheet.cell(row_index, 3).value)
        channel_type = normalize_text(worksheet.cell(row_index, 4).value)
        channels = read_animation_channels(worksheet, row_index, channel_headers, omit_zero=True)

        if mode_name is None and time_ms is None and channel_type is None and not channels:
            finalize_animation()
            continue
        if channel_type == "K factory":
            continue

        if mode_name is None or time_ms is None or channel_type is None:
            raise ValueError(f"Animation_Cfg 第 {row_index} 行缺少 MODE/Times(ms)/CHANNEL，无法解析。")

        if (
            current_animation is None
            or current_animation["mode_name"] != mode_name
            or current_animation["channel_type"] != channel_type
        ):
            finalize_animation()
            current_animation = {
                "animation_id": len(animations),
                "animation_kind": classify_animation_kind(mode_name),
                "mode_name": mode_name,
                "channel_type": channel_type,
                "source_row_start": row_index,
                "source_row_end": row_index,
                "frames": [],
            }

        current_animation["frames"].append(
            {
                "frame_id": len(current_animation["frames"]),
                "source_row": row_index,
                "time_ms": time_ms,
                "channels": channels,
            }
        )
        current_animation["source_row_end"] = row_index

    finalize_animation()
    animation_counts = summarize_animation_kind_counts(animations)

    result = {
        "parser": "animation_cfg",
        "sheet_name_raw": worksheet.title,
        "sheet_name": worksheet.title.strip(),
        "total_ic_count": total_ic_count,
        "channel_count_per_ic": channel_count_per_ic,
        **animation_counts,
        "header": header,
        "channel_headers": [channel_name for _, channel_name in channel_headers],
        "animations": animations,
    }
    return result


def parse_ti_sequential(worksheet: Worksheet) -> dict[str, Any]:
    header_row_index = next(
        (
            row_index
            for row_index in range(1, worksheet.max_row + 1)
            if normalize_text(worksheet.cell(row_index, 1).value) == "Times(ms)"
            and normalize_text(worksheet.cell(row_index, 2).value) == "CHANNEL"
        ),
        None,
    )
    if header_row_index is None:
        raise ValueError("TI_sequential 中未找到 Times(ms)/CHANNEL 表头。")

    channel_headers = find_animation_channel_headers(worksheet, header_row_index, start_column=3)
    total_ic_count, channel_count_per_ic = summarize_animation_channels(channel_headers)
    header = {
        "time_ms": normalize_text(worksheet.cell(header_row_index, 1).value),
        "channel_type": normalize_text(worksheet.cell(header_row_index, 2).value),
    }

    animation: dict[str, Any] | None = None
    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        time_ms = normalize_json_scalar(worksheet.cell(row_index, 1).value)
        channel_type = normalize_text(worksheet.cell(row_index, 2).value)
        channels = read_animation_channels(worksheet, row_index, channel_headers, omit_zero=True)

        if time_ms is None and channel_type is None and not channels:
            continue
        if time_ms is None or channel_type is None:
            raise ValueError(f"TI_sequential 第 {row_index} 行缺少 Times(ms)/CHANNEL，无法解析。")

        if animation is None:
            animation = {
                "animation_id": 0,
                "channel_type": channel_type,
                "source_row_start": row_index,
                "source_row_end": row_index,
                "frames": [],
            }
        elif animation["channel_type"] != channel_type:
            raise ValueError(
                f"TI_sequential 第 {row_index} 行 CHANNEL 类型为 {channel_type!r}，"
                f"与前面帧的 {animation['channel_type']!r} 不一致。"
            )

        animation["frames"].append(
            {
                "frame_id": len(animation["frames"]),
                "source_row": row_index,
                "time_ms": time_ms,
                "channels": channels,
            }
        )
        animation["source_row_end"] = row_index

    if animation is None:
        raise ValueError("TI_sequential 中未找到动画帧。")

    return {
        "parser": "ti_sequential",
        "sheet_name_raw": worksheet.title,
        "sheet_name": worksheet.title.strip(),
        "total_ic_count": total_ic_count,
        "channel_count_per_ic": channel_count_per_ic,
        "header": header,
        "channel_headers": [channel_name for _, channel_name in channel_headers],
        "animation": animation,
    }


def extract_formula_text(worksheet: Worksheet, row_index: int, column_index: int) -> str | None:
    value = worksheet.cell(row_index, column_index).value
    if not isinstance(value, str):
        return None
    text = value.replace("\r\n", "\n").strip()
    if not text.startswith("="):
        return None
    return text


def resolve_sheet_value(
    worksheet: Worksheet,
    value_worksheet: Worksheet,
    row_index: int,
    column_index: int,
) -> Any:
    resolved = normalize_json_scalar(value_worksheet.cell(row_index, column_index).value)
    if resolved is not None:
        return resolved
    return normalize_json_scalar(worksheet.cell(row_index, column_index).value)


def parse_named_value_rows(
    worksheet: Worksheet,
    value_worksheet: Worksheet,
    row_indices: range,
    *,
    name_column: int,
    value_column: int,
) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for row_index in row_indices:
        name = normalize_text(worksheet.cell(row_index, name_column).value)
        if name is None:
            continue
        entry = {
            "source_row": row_index,
            "name": name,
            "value": resolve_sheet_value(worksheet, value_worksheet, row_index, value_column),
        }
        formula = extract_formula_text(worksheet, row_index, value_column)
        if formula is not None:
            entry["value_formula"] = formula
        entries.append(entry)
    return entries


def condense_named_value_rows(entries: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        entry["name"]: entry["value"]
        for entry in entries
    }


def parse_position_row(
    worksheet: Worksheet,
    value_worksheet: Worksheet,
    row_index: int,
) -> dict[str, Any]:
    position = {
        "source_row": row_index,
        "position_name": normalize_text(worksheet.cell(row_index, 2).value),
        "steps_to_pos1_fs": resolve_sheet_value(worksheet, value_worksheet, row_index, 3),
        "motor_head_spindle_distance_to_pos1_mm": resolve_sheet_value(worksheet, value_worksheet, row_index, 4),
        "position_on_wall_mm": resolve_sheet_value(worksheet, value_worksheet, row_index, 5),
        "angle": resolve_sheet_value(worksheet, value_worksheet, row_index, 6),
    }
    for column_index, key in (
        (3, "steps_to_pos1_fs"),
        (4, "motor_head_spindle_distance_to_pos1_mm"),
        (5, "position_on_wall_mm"),
        (6, "angle"),
    ):
        formula = extract_formula_text(worksheet, row_index, column_index)
        if formula is not None:
            position[f"{key}_formula"] = formula
    return position


def parse_motor_control_modes(
    worksheet: Worksheet,
    value_worksheet: Worksheet,
    row_indices: range,
) -> list[dict[str, Any]]:
    control_modes: list[dict[str, Any]] = []
    current_mode: dict[str, Any] | None = None
    for row_index in row_indices:
        mode_name = normalize_text(worksheet.cell(row_index, 2).value)
        parameter_name = normalize_text(worksheet.cell(row_index, 3).value)
        if mode_name is None and parameter_name is None:
            continue

        if mode_name is not None:
            if current_mode is not None:
                control_modes.append(current_mode)
            current_mode = {
                "mode_name": mode_name,
                "source_row_start": row_index,
                "source_row_end": row_index,
                "parameters": [],
            }

        if current_mode is None:
            raise ValueError(f"Motor_Cfg 第 {row_index} 行缺少控制模式名称，无法解析。")
        if parameter_name is None:
            continue

        parameter = {
            "source_row": row_index,
            "name": parameter_name,
            "value": resolve_sheet_value(worksheet, value_worksheet, row_index, 6),
        }
        formula = extract_formula_text(worksheet, row_index, 6)
        if formula is not None:
            parameter["value_formula"] = formula
        current_mode["parameters"].append(parameter)
        current_mode["source_row_end"] = row_index

    if current_mode is not None:
        control_modes.append(current_mode)
    return control_modes


def condense_motor_control_modes(control_modes: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {
        mode["mode_name"]: {
            parameter["name"]: parameter["value"]
            for parameter in mode["parameters"]
        }
        for mode in control_modes
    }


MOTOR_CFG_SAFETY_KEY_MAP = {
    "StepperMotor_Lowvoltage(V)": "low_voltage_v",
    "StepperMotor_Overvoltage(V)": "over_voltage_v",
}
MOTOR_CFG_GENERAL_KEY_MAP = {
    "Stepper Motor Action For Positive Command": "positive_command_action",
    "Full step for motor head spindle moves 1mm": "full_steps_per_mm",
    "Distance ratio(command value on 10m wall/motor head spindle moves 1mm)": "wall_command_ratio_per_mm_at_10m",
}
MOTOR_CFG_CONTROL_MODE_KEY_MAP = {
    "Reference Run": "reference_run",
    "Normal Run": "normal_run",
}
MOTOR_CFG_CONTROL_PARAM_KEY_MAP = {
    "Runnning Current": "running_current",
    "Holding Current": "holding_current",
    "MAX Acceleration Ramp": "max_acceleration_ramp",
    "Minimum Movement Speed": "min_speed",
    "Normal Movement speed": "normal_speed",
    "Maximum Movement Speed": "max_speed",
}


def slug_ascii_key(text: str) -> str:
    return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", text.strip().lower())).strip("_")


def map_motor_cfg_key(name: str, key_map: dict[str, str]) -> str:
    return key_map.get(name, slug_ascii_key(name))


def split_motor_position_name(position_name: str) -> tuple[str, str]:
    match = re.fullmatch(r"(Pos\d+)\((.*)\)", position_name.strip())
    if match is None:
        return slug_ascii_key(position_name), position_name
    return match.group(1).lower(), match.group(2).strip()


def motor_afs_level_key(level_name: str) -> str:
    return slug_ascii_key(level_name.removesuffix(" position"))


def parse_motor_cfg(worksheet: Worksheet, value_worksheet: Worksheet) -> dict[str, Any]:
    title = normalize_text(worksheet.cell(1, 1).value)
    if title is None:
        raise ValueError("Motor_Cfg 中未找到标题。")

    position_headers = {
        "position_name": normalize_text(worksheet.cell(8, 2).value),
        "steps_to_pos1_fs": normalize_text(worksheet.cell(8, 3).value),
        "motor_head_spindle_distance_to_pos1_mm": normalize_text(worksheet.cell(8, 4).value),
        "position_on_wall_mm": normalize_text(worksheet.cell(8, 5).value),
        "angle": normalize_text(worksheet.cell(8, 6).value),
    }
    afs_headers = {
        "c_mode_position": normalize_text(worksheet.cell(28, 4).value),
        "v_mode_position": normalize_text(worksheet.cell(28, 5).value),
        "e_mode_position": normalize_text(worksheet.cell(28, 6).value),
    }

    positions = [parse_position_row(worksheet, value_worksheet, row_index) for row_index in range(9, 14)]
    afs_levels = [
        {
            "source_row": row_index,
            "level_name": normalize_text(worksheet.cell(row_index, 1).value),
            "c_mode_position": resolve_sheet_value(worksheet, value_worksheet, row_index, 4),
            "v_mode_position": resolve_sheet_value(worksheet, value_worksheet, row_index, 5),
            "e_mode_position": resolve_sheet_value(worksheet, value_worksheet, row_index, 6),
        }
        for row_index in range(29, 33)
        if normalize_text(worksheet.cell(row_index, 1).value) is not None
    ]

    return {
        "parser": "motor_cfg",
        "sheet_name_raw": worksheet.title,
        "sheet_name": worksheet.title.strip(),
        "title": title,
        "safety_voltage_configuration": parse_named_value_rows(
            worksheet,
            value_worksheet,
            range(3, 5),
            name_column=2,
            value_column=6,
        ),
        "general": parse_named_value_rows(
            worksheet,
            value_worksheet,
            range(5, 8),
            name_column=2,
            value_column=6,
        ),
        "position_headers": position_headers,
        "positions": positions,
        "control_modes": parse_motor_control_modes(worksheet, value_worksheet, range(14, 26)),
        "step_mode": {
            "source_row": 26,
            "label": normalize_text(worksheet.cell(26, 1).value),
            "value": resolve_sheet_value(worksheet, value_worksheet, 26, 4),
        },
        "afs_headers": afs_headers,
        "afs_levels": afs_levels,
    }


def relabel_values(field_labels: dict[str, str], values: dict[str, str]) -> dict[str, str]:
    del field_labels
    return {
        OUTPUT_VALUE_KEYS[key]: clean_output_value(value)
        for key, value in values.items()
        if value is not None
    }


def option_to_value_row(item: dict[str, Any], option: dict[str, Any]) -> dict[str, str]:
    raw_values = {key: option[key] for key in FIELD_LABELS if option.get(key) is not None}
    return relabel_values(item.get("field_labels", {}), raw_values)


def condense_hcm_prilin_matrix(parsed: dict[str, Any]) -> dict[str, Any]:
    condensed_items: list[dict[str, Any]] = []
    for item in parsed["items"]:
        condensed_item = {
            "id": item["id"],
            "name": item["name"],
        }

        defaults = item.get("defaults", {})
        options = item.get("options", [])
        if defaults:
            condensed_item["values"] = relabel_values(item.get("field_labels", {}), defaults)
        elif options:
            condensed_item["values"] = [option_to_value_row(item, option) for option in options]
        else:
            condensed_item["values"] = {}

        condensed_items.append(condensed_item)

    return {
        "sheet_name": parsed["sheet_name"],
        "items": condensed_items,
    }


def condense_ch_cfg(parsed: dict[str, Any]) -> dict[str, Any]:
    return {
        "sheet_name": parsed["sheet_name"],
        "config_type_descriptions": parsed["config_type_descriptions"],
        "ics": [
            {
                "ic_id": ic["ic_id"],
                "ic_name": ic["ic_name"],
                "channels": ic["channels"],
            }
            for ic in parsed["ics"]
        ],
    }


def condense_animation_cfg(parsed: dict[str, Any]) -> dict[str, Any]:
    grouped_animations = group_animations_by_kind(parsed["animations"])

    def simplify(animation: dict[str, Any]) -> dict[str, Any]:
        return {
            "mode_name": animation["mode_name"],
            "channel_type": animation["channel_type"],
            "frames": [
                {
                    "time_ms": frame["time_ms"],
                    "channels": frame["channels"],
                }
                for frame in animation["frames"]
            ],
        }

    result = {
        "sheet_name": parsed["sheet_name"],
        "total_ic_count": parsed["total_ic_count"],
        "channel_count_per_ic": parsed["channel_count_per_ic"],
        "unlock_animation_count": parsed["unlock_animation_count"],
        "lock_animation_count": parsed["lock_animation_count"],
        "unlock_animations": [simplify(animation) for animation in grouped_animations["unlock"]],
        "lock_animations": [simplify(animation) for animation in grouped_animations["lock"]],
    }
    if grouped_animations["other"]:
        result["other_animations"] = [simplify(animation) for animation in grouped_animations["other"]]
    return result


def condense_ti_sequential(parsed: dict[str, Any]) -> dict[str, Any]:
    return {
        "sheet_name": parsed["sheet_name"],
        "total_ic_count": parsed["total_ic_count"],
        "channel_count_per_ic": parsed["channel_count_per_ic"],
        "animation": {
            "channel_type": parsed["animation"]["channel_type"],
            "frames": [
                {
                    "time_ms": frame["time_ms"],
                    "channels": frame["channels"],
                }
                for frame in parsed["animation"]["frames"]
            ],
        },
    }


def condense_motor_cfg(parsed: dict[str, Any]) -> dict[str, Any]:
    motor_config = {
        "safety_voltage": {
            map_motor_cfg_key(entry["name"], MOTOR_CFG_SAFETY_KEY_MAP): entry["value"]
            for entry in parsed["safety_voltage_configuration"]
        },
        "general_settings": {
            map_motor_cfg_key(entry["name"], MOTOR_CFG_GENERAL_KEY_MAP): entry["value"]
            for entry in parsed["general"]
        },
        "control_modes": {
            map_motor_cfg_key(mode["mode_name"], MOTOR_CFG_CONTROL_MODE_KEY_MAP): {
                map_motor_cfg_key(parameter["name"], MOTOR_CFG_CONTROL_PARAM_KEY_MAP): parameter["value"]
                for parameter in mode["parameters"]
            }
            for mode in parsed["control_modes"]
        },
        "microstep_mode": parsed["step_mode"]["value"],
        "positions": {},
        "afs_positions": {},
    }

    for position in parsed["positions"]:
        position_key, label = split_motor_position_name(position["position_name"])
        motor_config["positions"][position_key] = {
            "label": label,
            "steps_to_pos1_fs": position["steps_to_pos1_fs"],
            "spindle_distance_to_pos1_mm": position["motor_head_spindle_distance_to_pos1_mm"],
            "wall_position_mm": position["position_on_wall_mm"],
            "angle_deg": position["angle"],
        }

    for level in parsed["afs_levels"]:
        motor_config["afs_positions"][motor_afs_level_key(level["level_name"])] = {
            "c_mode": level["c_mode_position"],
            "v_mode": level["v_mode_position"],
            "e_mode": level["e_mode_position"],
        }

    return {
        "sheet_name": parsed["sheet_name"],
        "title": parsed["title"],
        "motor_config": motor_config,
    }


def parse_sheet(worksheet: Worksheet, value_worksheet: Worksheet | None = None) -> dict[str, Any]:
    sheet_name = worksheet.title.strip()
    if sheet_name == "HCM_PriLIN_Matrix":
        return parse_hcm_prilin_matrix(worksheet)
    if sheet_name == "CH_Cfg":
        return parse_ch_cfg(worksheet)
    if sheet_name == "Animation_Cfg":
        return parse_animation_cfg(worksheet)
    if sheet_name == "TI_sequential":
        return parse_ti_sequential(worksheet)
    if sheet_name == "Motor_Cfg":
        return parse_motor_cfg(worksheet, value_worksheet or worksheet)
    raise ValueError(f"当前暂不支持 sheet: {worksheet.title!r}")


def condense_sheet(parsed: dict[str, Any]) -> dict[str, Any]:
    parser_name = parsed["parser"]
    if parser_name == "hcm_prilin_matrix":
        return condense_hcm_prilin_matrix(parsed)
    if parser_name == "ch_cfg":
        return condense_ch_cfg(parsed)
    if parser_name == "animation_cfg":
        return condense_animation_cfg(parsed)
    if parser_name == "ti_sequential":
        return condense_ti_sequential(parsed)
    if parser_name == "motor_cfg":
        return condense_motor_cfg(parsed)
    raise ValueError(f"当前暂不支持 parser: {parser_name}")


def resolve_output_path(output_path: Path | None, sheet_name: str, multi_sheet: bool = False) -> Path:
    if output_path is None:
        return OUTPUT_DIR / f"{sheet_name}.json"
    if not multi_sheet:
        return output_path
    if output_path.exists() and output_path.is_file():
        raise ValueError("未指定 --sheet 时，--output 需要是目录路径，不能是文件。")
    if not output_path.exists() and output_path.suffix.lower() == ".json":
        raise ValueError("未指定 --sheet 时，--output 需要是目录路径，不能是单个 JSON 文件。")
    return output_path / f"{sheet_name}.json"


def find_sheet(workbook, requested_sheet: str) -> Worksheet:
    matches = [sheet for sheet in workbook.worksheets if sheet.title.strip() == requested_sheet.strip()]
    if not matches:
        available = ", ".join(repr(sheet.title) for sheet in workbook.worksheets)
        raise ValueError(f"未找到 sheet: {requested_sheet!r}。可用 sheet: {available}")
    if len(matches) > 1:
        names = ", ".join(repr(sheet.title) for sheet in matches)
        raise ValueError(f"去空格后存在重名 sheet: {names}")
    return matches[0]


def find_supported_sheets(workbook) -> list[Worksheet]:
    supported_sheets: list[Worksheet] = []
    seen: dict[str, Worksheet] = {}
    for sheet in workbook.worksheets:
        sheet_name = sheet.title.strip()
        if sheet_name not in SUPPORTED_SHEET_NAMES:
            continue
        if sheet_name in seen:
            names = ", ".join(repr(item.title) for item in (seen[sheet_name], sheet))
            raise ValueError(f"去空格后存在重名 sheet: {names}")
        seen[sheet_name] = sheet
        supported_sheets.append(sheet)

    if not supported_sheets:
        available = ", ".join(repr(sheet.title) for sheet in workbook.worksheets)
        supported = ", ".join(repr(sheet_name) for sheet_name in SUPPORTED_SHEET_NAMES)
        raise ValueError(f"未找到任何已支持的 sheet。已支持 sheet: {supported}。可用 sheet: {available}")
    return supported_sheets


def count_output_items(output_payload: dict[str, Any], parsed: dict[str, Any]) -> int:
    if "items" in output_payload:
        return len(output_payload["items"])
    if "ics" in output_payload:
        return len(output_payload["ics"])
    if "animations" in output_payload:
        return sum(len(animation.get("frames", [])) for animation in output_payload["animations"])
    if "unlock_animations" in output_payload or "lock_animations" in output_payload:
        return sum(
            len(animation.get("frames", []))
            for key in ("unlock_animations", "lock_animations", "other_animations")
            for animation in output_payload.get(key, [])
        )
    if "motor_config" in output_payload:
        motor_config = output_payload["motor_config"]
        return (
            len(motor_config.get("safety_voltage", {}))
            + len(motor_config.get("general_settings", {}))
            + len(motor_config.get("positions", {}))
            + sum(len(parameters) for parameters in motor_config.get("control_modes", {}).values())
            + (1 if motor_config.get("microstep_mode") is not None else 0)
            + len(motor_config.get("afs_positions", {}))
        )
    if "animation" in output_payload:
        return len(output_payload["animation"].get("frames", []))
    if "positions" in output_payload and "control" in output_payload and "afs_levels" in output_payload:
        return (
            len(output_payload.get("safety_voltage_configuration", {}))
            + len(output_payload.get("general", {}))
            + len(output_payload.get("positions", []))
            + sum(len(parameters) for parameters in output_payload.get("control", {}).values())
            + (1 if output_payload.get("step_mode") is not None else 0)
            + len(output_payload.get("afs_levels", []))
        )
    if "items" in parsed:
        return len(parsed["items"])
    if "ics" in parsed:
        return len(parsed["ics"])
    if "animations" in parsed:
        return sum(len(animation.get("frames", [])) for animation in parsed["animations"])
    if "animation" in parsed:
        return len(parsed["animation"].get("frames", []))
    if "positions" in parsed and "control_modes" in parsed and "afs_levels" in parsed:
        return (
            len(parsed.get("safety_voltage_configuration", []))
            + len(parsed.get("general", []))
            + len(parsed.get("positions", []))
            + sum(len(mode.get("parameters", [])) for mode in parsed.get("control_modes", []))
            + (1 if parsed.get("step_mode") else 0)
            + len(parsed.get("afs_levels", []))
        )
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Convert CHCM workbook sheets to JSON.")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=None,
        help="Workbook path. Overrides the value loaded from .config and Kconfig.",
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=KCONFIG_CONFIG_DEFAULT,
        help=f"Kconfig .config path. Default: {KCONFIG_CONFIG_DEFAULT}",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name. Leading and trailing spaces are ignored when matching. If omitted, converts all supported sheets.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output JSON path for a single sheet, or output directory when converting multiple sheets. Default: output/<sheet_name>.json",
    )
    parser.add_argument(
        "--mode",
        choices=("values", "full"),
        default="values",
        help="Output only config items and values, or the full parsed structure.",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    workbook_path = resolve_workbook_path(args.workbook, args.config)
    workbook = load_workbook(workbook_path, data_only=False)
    value_workbook = load_workbook(workbook_path, data_only=True)
    try:
        worksheets = [find_sheet(workbook, args.sheet)] if args.sheet else find_supported_sheets(workbook)
        batch_mode = args.sheet is None

        print(f"Workbook: {workbook_path}")
        print(f"Mode: {args.mode}")
        for worksheet in worksheets:
            value_worksheet = find_sheet(value_workbook, worksheet.title)
            parsed = parse_sheet(worksheet, value_worksheet)
            output_payload = condense_sheet(parsed) if args.mode == "values" else parsed
            output_path = resolve_output_path(args.output, parsed["sheet_name"], multi_sheet=batch_mode)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(json.dumps(output_payload, ensure_ascii=False, indent=2), encoding="utf-8")
            item_count = count_output_items(output_payload, parsed)

            print(f"Wrote {output_path}")
            print(f"Sheet: {parsed['sheet_name_raw']!r}")
            print(f"Items: {item_count}")
    finally:
        value_workbook.close()
        workbook.close()


if __name__ == "__main__":
    main()
